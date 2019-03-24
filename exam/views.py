from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http.response import HttpResponseRedirect
from django.shortcuts import render, redirect, reverse
from exam.forms import CreateUser, Login, NewExam, OptionSubmit, UploadFileForm
from exam.models import Exam, Question
from django.template.defaulttags import register

from openpyxl import load_workbook

@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)

@register.filter
def answer(id):
    pass


def welcome(request):

    return render(request, "homepage.html")

def create_user(request):
    if request.method == 'POST':
        form = CreateUser(request.POST)
        if form.is_valid():
            username = form.cleaned_data["username"]
            email = form.cleaned_data["email"]
            password = form.cleaned_data["password"]
            user = User.objects.create_user(username=username,
                                            email=email,
                                            password=password)
            user.save()
            login(request, user)

            return redirect("welcome")


    else:
        form = CreateUser()
        context = {
            "form" : form,
        }

    return render(request, "createuser.html", context)

def log_in(request):
    form = Login()
    context = {
        "form" : form,
    }
    if request.method == 'POST':
        form = Login(request.POST)
        if form.is_valid():
            username = request.POST['username']
            password = request.POST['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("welcome")
            else:
                error_message = "Your password/username did not match. Please try again."
                context = {
                    "error_message" : error_message,
                    "form" : form,
                }
                return render(request, "login.html", context)

    return render(request, "login.html", context)

@login_required()
def new_exam(request):
    form = NewExam()
    context = {
        "form": form,
    }
    if request.method == 'POST':
        form = NewExam(request.POST)
        if form.is_valid():
            subject = request.POST['subject']
            question_amount = request.POST['question_amount']
            answer_at_end = request.POST['feedback']
            e = Exam(subject=subject, user=request.user, question_amount=question_amount, answer_at_end=answer_at_end)
            e.save()
            e.set_questions()
            return HttpResponseRedirect(reverse('question_form', kwargs={'pk': e.id}))

        else:
            return redirect("exam_templates/new_exam.html", context)

    return render(request, "exam_templates/newexam.html", context)

@login_required
def previous_exams(request):

    all_exams = Exam.objects.filter(user=request.user).order_by("-id")
    context = {
        "all_exams" : all_exams,
    }
    return render(request, "exam_templates/previousexams.html", context)

def question_form(request, pk):
    current_exam = Exam.objects.get(id=int(pk))

    try:
        question_object = current_exam.get_question()
    except:
        current_exam.is_complete = True
        current_exam.save()
        return HttpResponseRedirect(reverse('results', kwargs={'pk': pk}))

    form = OptionSubmit(question_object)
    context = {
        "form" : form,
        "pk" : pk,
        "exam" : current_exam,
    }
    if request.method == 'POST':
        form = OptionSubmit(question_object, request.POST)
        if form.is_valid():
            submitted_answer = request.POST['submitted_answer']
            if current_exam.answer_at_end is True:              ## if right and answer at end
                if submitted_answer == question_object.answer:
                    current_exam.current_question += 1
                    current_exam.score += 1
                    current_exam.save()
                    return HttpResponseRedirect(reverse('question_form', kwargs={'pk': pk}))
                else:                                           ### if wrong and answer at end
                    current_exam.current_question += 1
                    incorrect_list = [current_exam.incorrect_questions]
                    incorrect_list.append(str(question_object.id))
                    incorrect_answers_list = [current_exam.incorrect_answers]
                    incorrect_answers_list.append(str(submitted_answer))
                    current_exam.incorrect_questions = ",".join(incorrect_list)
                    current_exam.incorrect_answers = ",".join(incorrect_answers_list)
                    current_exam.save()
                    return HttpResponseRedirect(reverse('question_form', kwargs={'pk': pk}))
            else:
                if submitted_answer == question_object.answer:   ### if right and answer at question
                    current_exam.current_question += 1
                    current_exam.score += 1
                    current_exam.save()
                    options = {
                        "a": question_object.choiceA,
                        "b": question_object.choiceB,
                        "c": question_object.choiceC,
                        "d": question_object.choiceD,
                    }
                    context = {
                        "message": "Correct Answer",
                        "question": question_object,
                        "selected_answer": submitted_answer,
                        "pk": pk,
                        "options": options,
                        "exam" : current_exam,
                    }
                    return render(request, "exam_templates/questionform.html", context)

                else:                                          ### if wrong and answer at question
                    current_exam.current_question += 1
                    incorrect_list = [current_exam.incorrect_questions]
                    incorrect_list.append(str(question_object.id))
                    incorrect_answers_list = [current_exam.incorrect_answers]
                    incorrect_answers_list.append(str(submitted_answer))
                    current_exam.incorrect_questions = ",".join(incorrect_list)
                    current_exam.incorrect_answers = ",".join(incorrect_answers_list)
                    current_exam.save()
                    options = {
                        "a": question_object.choiceA,
                        "b": question_object.choiceB,
                        "c": question_object.choiceC,
                        "d": question_object.choiceD,
                    }
                    context = {
                        "message" : question_object.answer_text,
                        "question" : question_object,
                        "selected_answer" : submitted_answer,
                        "pk" : pk,
                        "options" : options,
                        "exam" : current_exam,
                        "question_text" : question_object.text,
                    }
                    return render(request, "exam_templates/questionform.html", context)

    return render(request, "exam_templates/questionform.html", context)

@login_required()
def results(request, pk):
    current_exam = Exam.objects.get(id=int(pk))
    incorrect_questions = current_exam.get_incorrect_questions()
    incorrect_answers = current_exam.get_incorrect_answers()
    result_message = current_exam.result_message()
    percent = current_exam.result_percentage()
    answer_question_list = list(zip(incorrect_questions, incorrect_answers))
    print(list(answer_question_list))
    context = {
        "exam" : current_exam,
        "pk" : pk,
        "incorrect_questions": incorrect_questions,
        "incorrect_answers": incorrect_answers,
        "result_message" : result_message,
        "percent" : percent,
        "answer_question_list" : answer_question_list,
    }

    return render(request, "exam_templates/results.html", context)

def add_question(request):
    pass

def log_out(request):
    logout(request)
    return redirect("welcome")

def testpage(request):
    return render(request, "index.html")

def upload_file(request):
    def to_database(file):
        book = load_workbook(file)
        sheet = book.active

        for i in range(2, 191):
            ques_obj = Question.objects.create(
                text=sheet.cell(row=i, column=4).value,
                topic = sheet.cell(row=i, column=3).value,
                choiceA = sheet.cell(row=i, column=5).value,
                choiceB = sheet.cell(row=i, column=6).value,
                choiceC = sheet.cell(row=i, column=7).value,
                choiceD = sheet.cell(row=i, column=8).value,
                answer = sheet.cell(row=i, column=10).value,
                answer_text = sheet.cell(row=i, column=11).value,
                subject = sheet.cell(row=i, column=2).value,
                context = sheet.cell(row=i, column=9).value
                )
            ques_obj.save()
            ques_obj.id = (i - 1)
            ques_obj.save()

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            to_database(request.FILES['file'])
            return HttpResponseRedirect('welcome.html')
    else:
        form = UploadFileForm()
    return render(request, 'exam_templates/upload.html', {'form': form})



